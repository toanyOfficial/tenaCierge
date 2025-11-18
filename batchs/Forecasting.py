#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
================================================================================
Forecasting 프로그램 명세서 v1.3  +  Debug Points  [주석/코드 포함]
================================================================================

1) 기본 개념
- data.toml을 기반으로 각 객실의 ICS 캘린더를 다운로드하고, 특정 날짜의
  확정 퇴실(out)과 예측 퇴실(potential)을 계산한다.
- ICS의 DTEND로 확정 out을 판단하고, 모델 파라미터로 p_out(퇴실 확률)을 추정.
- 모델은 D−1(단기)과 D−7(주간) 두 세트를 유지. D−2~D−6은 선형보간, 7일 이상은 D−7 세트 적용.
- 튜닝은 p_out 기반(Brier Score)으로 수행. 컷오프(high)는 precision/업무량 기준으로 조정.

2) 날짜 입력 및 실행
- 형식: YYYYMMDD (중간 구분자 없음)
- 입력: 콤마(,) 다중, 틸다(~) 범위(예: 20251108,20251110~20251112)
- --auto: D0 기준 D+1~D+8 자동 생성(매일 16:00 가정)
- 입력 오류 시 재입력

3) 파일명 및 출력
- forecasting(YYYYMMDD).xlsx (가장 이른 날짜 기준, 중복 시 (1),(2) 자동증가)
- 출력 구조:
  forecasting(YYYYMMDD).xlsx     # 결과 파일(날짜별 시트 + Summary)
  ics/YYYYMMDDhhmmss/            # ICS 다운로드 폴더
  model_state.toml               # 모델 변수(D1/D7, threshold) 통합 저장
  report.xlsx                    # 리포트(시트 4개: D-7, D-1, Accuracy, Tuning)

4) ICS 다운로드
- data.toml의 url/urls/ics 필드를 그대로 사용(정규식 파싱 금지)
- 저장 경로: ./ics/YYYYMMDDhhmmss/
- 파일명: [name]_YYYYMMDDhhmmss_[platform].ics
  * name: data.toml의 name 또는 room
  * platform: toml에 있으면 사용, 없으면 URL로 유추(airbnb→air, booking→booking)
- 404/네트워크 오류는 해당 방만 스킵, 진행 로그 출력

5) 시트 구성
- 날짜별 시트: Header(건물 요약) + Detail(객실 상세)
- Summary: 날짜별 TOTAL 행(검정배경+흰글씨+bold) + 건물별 행

6) Header 컬럼
- sector | building | out | early | p50 | low | hi | potential | total

7) Detail 컬럼
- sector | building | room | out | p_out | potential | total
  * out: 확정 퇴실 시각(HH:MM, 없으면 공백)
  * p_out: 예측 확률(0.00~1.00)
  * potential: ○(≥ high), △(≥ borderline), 공백
  * total: out + potential(○)

8) 정렬
- Header: sector, out DESC
- Detail: sector, out 유무, building, room

9) 확정 로직 (out/early)
- DTEND ∈ [해당일 00:00, 익일 00:00) → out
- out_time < 12:00 → early
- back-to-back(end == next.start)은 병합하지 않음(중간 out 발생)
- overlap(next.start < prev.end)은 병합(연속투숙)

10) 예측 로직 (Potential)
- 기본식: p_out = sigmoid(alpha + beta * weekday_score)
- 보정: D+1 → ×0.6,  D+7 이상 → ×1.1,  요일계수 ×weekday_factor
- 분류: p_out≥high→○,  borderline≤p_out<high→△,  그 외 공백
- 집계: potential은 ○만 카운트, total = out + potential(○)

11) p50 / low / hi
- μ = Σp,  σ² = Σ p(1−p)
- p50=round(μ),  low=μ−1.28σ,  hi=μ+1.28σ
- potential=○만 카운트, total = out + potential(○)

12) report.xlsx (시트 4개)
- D-7: run_date | target_date | sector | building | room | p_out | pred_label | potential | actual_out | correct
- D-1: (동일, 단 라벨 기준은 d1_high)
- Accuracy: date | horizon(D-1/D-7) | acc | prec | rec | f1 | n
- Tuning: date | horizon | variable | before | after | delta | explanation
* D−2~D−6 예측은 report에 저장하지 않음(집중: D-1/D-7).
* D0 실행 시 D-1/D-7 시트의 target_date==오늘인 행에 actual_out/correct 채우고, Accuracy/Tuning 갱신.

13) model_state.toml
[threshold]
d1_high = 0.65
d7_high = 0.68
borderline = 0.40

[calibration]
d1_alpha = 0.12
d1_beta  = 0.94
d7_alpha = 0.15
d7_beta  = 1.02

- 없으면 자동 생성.
- 파라미터 적용 규칙:
  horizon=1 → D−1 세트
  horizon=2~6 → 선형보간
  horizon≥7 → D−7 세트
  ratio=(h-1)/6;  param=(1−ratio)*D1 + ratio*D7

14) 정확도 & 튜닝 (v1.3 변경점)
- 정확도: D0에서 실제 out 반영 → D-1/D-7 비교 → Accuracy 시트 기록
- 튜닝:
  * α,β는 p_out 기반 Brier Score 최소화(주로 D-7에 적용)
  * high는 precision(또는 업무량) 목표로 미세 조정(주로 D-1에 적용)
- Tuning 시트에 변수/이유 기록, model_state.toml 동기화

15) 종료
- --auto: 완료 후 CMD까지 즉시 종료(os._exit(0))
- 수동: 에러 시 10분 입력 대기 후 종료

----------------------------------- Debug Points --------------------------------
- URL: TOML 문자열 그대로 사용(정규식 추출 금지), 따옴표로 인용
- ICS 실패: 404/네트워크 오류는 해당 방만 스킵, 로그 남김
- 이벤트 병합: overlap만 병합(start<last.end), back-to-back은 분리
- out 판정: end ∈ [D00:00, D+1 00:00)
- p50/low/hi: μ, σ²=Σp(1−p) 기반 정규 근사
- 파라미터 저장: model_state.toml 하나로 통합(D1/D7)
- report.xlsx: D-7, D-1, Accuracy, Tuning 4시트 고정
- Auto 종료: os._exit(0)
- 정렬/서식: Header(sector, −out), Summary TOTAL 서식(검정/흰글씨/bold)
- horizon 적용: 1→D1, 2~6→보간, 7+→D7
- Accuracy/Tuning 타이밍: D0에서 actual 매칭→Accuracy→Tuning→state 저장
- 튜닝 기준: pred_label이 아니라 p_out 기반(Brier) + high는 precision/업무량으로 조정
================================================================================
"""

from __future__ import annotations
import argparse
import datetime as dt
import math
import os
import re
import sys
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional

import toml
import requests
from icalendar import Calendar
from dateutil import tz
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# -------------------- 전역 설정 --------------------
SEOUL = tz.gettz("Asia/Seoul")
TODAY = dt.datetime.now(SEOUL).date()

MODEL_STATE_PATH = "model_state.toml"
REPORT_XLSX = "report.xlsx"

# 요일 기본/계수 (가벼운 신호 + 미세 보정)
WEEKDAY_BASE = {0:0.6,1:0.45,2:0.45,3:0.5,4:0.8,5:1.0,6:0.9}
WEEKDAY_FACTOR = {0:0.95,1:0.90,2:0.90,3:0.95,4:1.00,5:1.05,6:1.00}

# D-1 precision 목표(임계값 튜닝용). 필요시 조정.
D1_PRECISION_TARGET = 0.70
D1_HIGH_STEP = 0.02
D1_HIGH_MIN, D1_HIGH_MAX = 0.40, 0.90

# -------------------- 데이터 구조 --------------------
@dataclass(frozen=True)
class RoomKey:
    sector: str
    building: str
    room: str

@dataclass
class Room:
    key: RoomKey
    out_time: dt.time
    urls: List[str]
    platform: Optional[str]  # 'air', 'booking', 기타/None

@dataclass
class Event:
    start: dt.datetime
    end: dt.datetime

# -------------------- 유틸 --------------------
def ensure_ics_dir(ts_str: str) -> str:
    ics_dir = os.path.join("ics", ts_str)
    os.makedirs(ics_dir, exist_ok=True)
    return ics_dir

def parse_hhmm(s: str) -> dt.time:
    h, m = s.split(":")
    return dt.time(int(h), int(m))

def to_aware(x) -> dt.datetime:
    if isinstance(x, dt.datetime):
        return x if x.tzinfo else x.replace(tzinfo=tz.UTC)
    if isinstance(x, dt.date):
        return dt.datetime(x.year, x.month, x.day, tzinfo=tz.UTC)
    raise TypeError("Unsupported date type")

def day_bounds(d: dt.date) -> Tuple[dt.datetime, dt.datetime]:
    s = dt.datetime(d.year, d.month, d.day, 0, 0, tzinfo=SEOUL)
    e = s + dt.timedelta(days=1)
    return s, e

def sigmoid(x: float) -> float:
    try:
        return 1.0 / (1.0 + math.exp(-x))
    except OverflowError:
        return 0.0 if x < 0 else 1.0

def infer_platform(url: str) -> str:
    u = url.lower()
    if "airbnb" in u:
        return "air"
    if "booking" in u:
        return "booking"
    return "other"

def unique_filename(base: str, ext: str) -> str:
    cand = f"{base}{ext}"
    if not os.path.exists(cand):
        return cand
    i = 1
    while True:
        cand = f"{base}({i}){ext}"
        if not os.path.exists(cand):
            return cand
        i += 1

# -------------------- data.toml 로딩 --------------------
def load_data_toml(path: str = "data.toml") -> List[Room]:
    data = toml.load(path)
    items = data.get("entries") or data.get("rooms") or []
    if not isinstance(items, list):
        raise RuntimeError("data.toml: 'rooms' 또는 'entries'는 리스트여야 합니다.")
    rooms: List[Room] = []
    seen = set()

    for it in items:
        sector = str(it.get("sector", "")).strip()
        building = str(it.get("building", "")).strip()
        name = str(it.get("name", it.get("room",""))).strip()
        out_s = str(it.get("out", "12:00")).strip()
        platform = it.get("platform", None)
        urls_field = it.get("url") or it.get("urls") or it.get("ics")

        if not (sector and building and name):
            raise RuntimeError(f"[에러] data.toml 필수 값 누락: {it}")

        urls: List[str] = []
        if isinstance(urls_field, str):
            if urls_field.strip():
                urls = [urls_field.strip()]
        elif isinstance(urls_field, list):
            urls = [str(x).strip() for x in urls_field if str(x).strip()]

        if platform is None and urls:
            platform = infer_platform(urls[0])

        key = RoomKey(sector, building, name)
        if key in seen:
            raise RuntimeError(f"[에러] 중복 방 발견: {sector}/{building}/{name}")
        seen.add(key)

        rooms.append(Room(
            key=key,
            out_time=parse_hhmm(out_s),
            urls=urls,
            platform=platform
        ))

    return rooms

# -------------------- ICS --------------------
def download_ics(rooms: List[Room], ts: str, ics_dir: str) -> Dict[RoomKey, List[str]]:
    session = requests.Session()
    session.headers.update({"User-Agent": "Forecasting/1.3"})
    total = sum(len(r.urls) for r in rooms)
    done = 0
    print(f"[ICS] 총 {total}개 다운로드 시작")

    mapping: Dict[RoomKey, List[str]] = {}
    for r in rooms:
        paths: List[str] = []
        if not r.urls:
            print(f"[경고] URL 없음: {r.key.sector}/{r.key.building}/{r.key.room}")
        for idx, url in enumerate(r.urls):
            fname = f"{r.key.room}_{ts}_{r.platform or 'other'}.ics"
            safe = re.sub(r"[^A-Za-z0-9_.\-]", "_", fname)
            out_path = os.path.join(ics_dir, safe if idx == 0 else safe.replace(".ics", f"_{idx}.ics"))
            try:
                resp = session.get(url, timeout=30)
                resp.raise_for_status()
                with open(out_path, "wb") as f:
                    f.write(resp.content)
                done += 1
                print(f"  [{done}/{total}] 저장: {out_path}")
                paths.append(out_path)
            except requests.exceptions.HTTPError as e:
                status = getattr(e.response, "status_code", None)
                if status == 404:
                    print(f"[경고] 404 Not Found: {url} → 스킵")
                else:
                    print(f"[경고] HTTP 오류: {url} ({e}) → 스킵")
            except Exception as e:
                print(f"[경고] 네트워크 오류: {url} ({e}) → 스킵")
        mapping[r.key] = paths
    print(f"[ICS] 다운로드 완료 ({done}/{total})")
    return mapping

def parse_ics(paths: List[str]) -> List[Event]:
    events: List[Event] = []
    for p in paths:
        try:
            with open(p, "rb") as f:
                cal = Calendar.from_ical(f.read())
            for comp in cal.walk("VEVENT"):
                dtstart = comp.get("DTSTART")
                dtend = comp.get("DTEND")
                if not (dtstart and dtend):
                    continue
                s = to_aware(dtstart.dt).astimezone(SEOUL)
                e = to_aware(dtend.dt).astimezone(SEOUL)
                events.append(Event(s, e))
        except Exception as e:
            print(f"[경고] ICS 파싱 실패: {p} ({e})")
    return merge_events(events)

def merge_events(events: List[Event]) -> List[Event]:
    if not events:
        return []
    evs = sorted(events, key=lambda x: x.start)
    merged = [Event(evs[0].start, evs[0].end)]
    for ev in evs[1:]:
        last = merged[-1]
        # overlap만 병합(start < last.end), back-to-back(start == last.end)은 분리
        if ev.start < last.end:
            last.end = max(last.end, ev.end)
        else:
            merged.append(Event(ev.start, ev.end))
    return merged

# -------------------- 비즈니스 로직 --------------------
def occupancy_flags(events: List[Event], d: dt.date) -> Dict[str, bool]:
    s, e = day_bounds(d)
    noon = dt.datetime(d.year, d.month, d.day, 12, 0, tzinfo=SEOUL)
    out_flag = any(s <= ev.end < e for ev in events)
    stay_flag = any(ev.start < noon < ev.end for ev in events)
    return {"out": out_flag, "stay": stay_flag}

def weekday_score(d: dt.date) -> float:
    return WEEKDAY_BASE[d.weekday()]

def horizon_params(h: int, state: dict) -> Tuple[float,float,float]:
    d1a = float(state["calibration"]["d1_alpha"])
    d1b = float(state["calibration"]["d1_beta"])
    d7a = float(state["calibration"]["d7_alpha"])
    d7b = float(state["calibration"]["d7_beta"])
    d1h = float(state["threshold"]["d1_high"])
    d7h = float(state["threshold"]["d7_high"])
    if h <= 1:
        return d1a, d1b, d1h
    if h >= 7:
        return d7a, d7b, d7h
    ratio = (h - 1) / 6.0
    alpha = (1 - ratio) * d1a + ratio * d7a
    beta  = (1 - ratio) * d1b + ratio * d7b
    high  = (1 - ratio) * d1h + ratio * d7h
    return alpha, beta, high

def p_room_out(events: List[Event], d: dt.date, h: int, state: dict) -> float:
    flags = occupancy_flags(events, d)
    # 확정/체류는 확률을 강제
    if flags["out"]:
        return 1.0
    if flags["stay"]:
        return 0.0
    alpha, beta, _ = horizon_params(h, state)
    p = sigmoid(alpha + beta * weekday_score(d))
    # horizon 보정
    if h == 1:
        p *= 0.6
    elif h >= 7:
        p *= 1.1
    p *= WEEKDAY_FACTOR[d.weekday()]
    return min(1.0, max(0.0, p))

def parse_dates(spec: str) -> List[dt.date]:
    parts = [p.strip() for p in spec.split(",") if p.strip()]
    out = set()
    for p in parts:
        if "~" in p:
            a, b = [x.strip() for x in p.split("~", 1)]
            d1 = dt.datetime.strptime(a, "%Y%m%d").date()
            d2 = dt.datetime.strptime(b, "%Y%m%d").date()
            if d2 < d1:
                raise ValueError(f"잘못된 범위: {p}")
            cur = d1
            while cur <= d2:
                out.add(cur)
                cur += dt.timedelta(days=1)
        else:
            out.add(dt.datetime.strptime(p, "%Y%m%d").date())
    return sorted(out)

# -------------------- 모델 상태 --------------------
def load_state() -> dict:
    if os.path.exists(MODEL_STATE_PATH):
        try:
            st = toml.load(MODEL_STATE_PATH)
            # 필수 키 보강
            st.setdefault("threshold", {})
            st["threshold"].setdefault("d1_high", 0.65)
            st["threshold"].setdefault("d7_high", 0.68)
            st["threshold"].setdefault("borderline", 0.40)
            st.setdefault("calibration", {})
            st["calibration"].setdefault("d1_alpha", 0.12)
            st["calibration"].setdefault("d1_beta", 0.94)
            st["calibration"].setdefault("d7_alpha", 0.15)
            st["calibration"].setdefault("d7_beta", 1.02)
            return st
        except Exception:
            pass
    st = {
        "threshold": {"d1_high":0.65, "d7_high":0.68, "borderline":0.40},
        "calibration": {"d1_alpha":0.12, "d1_beta":0.94, "d7_alpha":0.15, "d7_beta":1.02}
    }
    with open(MODEL_STATE_PATH, "w", encoding="utf-8") as f:
        toml.dump(st, f)
    return st

def save_state(st: dict):
    with open(MODEL_STATE_PATH, "w", encoding="utf-8") as f:
        toml.dump(st, f)

# -------------------- report.xlsx --------------------
def open_or_create_report() -> Workbook:
    if os.path.exists(REPORT_XLSX):
        try:
            return load_workbook(REPORT_XLSX)
        except Exception:
            pass
    wb = Workbook()
    ws = wb.active; ws.title = "D-7"
    ws.append(["run_date","target_date","sector","building","room","p_out","pred_label","potential","actual_out","correct"])
    ws2 = wb.create_sheet("D-1")
    ws2.append(["run_date","target_date","sector","building","room","p_out","pred_label","potential","actual_out","correct"])
    ws3 = wb.create_sheet("Accuracy")
    ws3.append(["date","horizon","acc","prec","rec","f1","n"])
    ws4 = wb.create_sheet("Tuning")
    ws4.append(["date","horizon","variable","before","after","delta","explanation"])
    wb.save(REPORT_XLSX)
    return wb

def append_prediction(wb: Workbook, sheet_name: str, rows: List[List]):
    ws = wb[sheet_name]
    for r in rows:
        ws.append(r)
    wb.save(REPORT_XLSX)

def _calc_metrics_for_sheet(wb: Workbook, sheet_name: str, today_s: str) -> Tuple[int,float,float,float,float]:
    ws = wb[sheet_name]
    tp=fp=fn=tn=0; n=0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] != today_s:  # target_date
            continue
        actual = row[8]; pred = row[6]
        if actual is None or pred is None:
            continue
        actual = int(actual); pred = int(pred)
        n += 1
        if pred==1 and actual==1: tp += 1
        elif pred==1 and actual==0: fp += 1
        elif pred==0 and actual==1: fn += 1
        else: tn += 1
    acc = (tp+tn)/n if n>0 else 0.0
    prec = tp/(tp+fp) if (tp+fp)>0 else 0.0
    rec = tp/(tp+fn) if (tp+fn)>0 else 0.0
    f1 = (2*prec*rec/(prec+rec)) if (prec+rec)>0 else 0.0
    return n, round(acc,3), round(prec,3), round(rec,3), round(f1,3)

def update_actuals_and_accuracy_and_tune(
    wb: Workbook, today: dt.date, rooms: List[Room], events_map: Dict[RoomKey, List[Event]], state: dict
):
    """
    D0:
      1) 오늘 실제 out 계산 → report D-1/D-7의 target_date==오늘 행에 actual_out/correct 채움
      2) Accuracy 시트 업데이트
      3) 튜닝:
         - D-7: p_out 기반 Brier Score로 d7_alpha/d7_beta 미세 조정
         - D-1: precision 목표로 d1_high 소폭 조정
      4) Tuning 시트 기록 + model_state.toml 저장
    """
    today_s = today.strftime("%Y%m%d")

    # 1) 오늘 실제 out 맵
    actual_map: Dict[Tuple[str,str,str], int] = {}
    for r in rooms:
        flags = occupancy_flags(events_map.get(r.key, []), today)
        actual_map[(r.key.sector, r.key.building, r.key.room)] = 1 if flags["out"] else 0

    def update_sheet(name: str):
        ws = wb[name]
        updated = 0
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[1].value != today_s:  # target_date
                continue
            key = (row[2].value, row[3].value, row[4].value)
            if key not in actual_map:
                continue
            actual = actual_map[key]
            row[8].value = actual  # actual_out
            pred_label = int(row[6].value or 0)
            row[9].value = 1 if pred_label == actual else 0  # correct
            updated += 1
        if updated:
            wb.save(REPORT_XLSX)

    # D-1, D-7 각각 actual 업데이트
    update_sheet("D-1")
    update_sheet("D-7")

    # 2) Accuracy 집계
    ws_acc = wb["Accuracy"]
    for sh, tag in [("D-1","D-1"),("D-7","D-7")]:
        n, acc, prec, rec, f1 = _calc_metrics_for_sheet(wb, sh, today_s)
        if n > 0:
            ws_acc.append([today_s, tag, acc, prec, rec, f1, n])
    wb.save(REPORT_XLSX)

    # 3) 튜닝
    ws_tune = wb["Tuning"]

    # 3-1) D-7: p_out 기반 Brier Score로 d7_alpha/d7_beta 미세 조정
    d7_ws = wb["D-7"]
    grads_alpha = []
    grads_beta  = []
    for row in d7_ws.iter_rows(min_row=2, values_only=True):
        if row[1] != today_s:  # target_date
            continue
        try:
            p = float(row[5])  # p_out
        except (TypeError, ValueError):
            continue
        y = row[8]
        if y is None:
            continue
        y = int(y)
        # L=(p-y)^2,  ∂L/∂p = 2(p-y) ;  p = σ(z)*C  (C: 보정계수 포함)
        # σ'(z)=p_raw(1-p_raw),  여기서는 근사적으로 p*(1-p) 사용(보정계수 영향 미세)
        dp = 2*(p - y)
        sigma_prime = p*(1-p)
        grad_common = dp * sigma_prime
        w = weekday_score(today)  # horizon 고정이 아니므로 대략 오늘 기준을 사용(간소화)
        grads_alpha.append(grad_common)
        grads_beta.append(grad_common * w)
    if grads_alpha:
        eta = 0.03  # 보수적 학습률
        before_a = float(state["calibration"]["d7_alpha"]); before_b = float(state["calibration"]["d7_beta"])
        after_a = before_a - eta * (sum(grads_alpha)/len(grads_alpha))
        after_b = before_b - eta * (sum(grads_beta)/len(grads_beta))
        state["calibration"]["d7_alpha"] = round(after_a, 4)
        state["calibration"]["d7_beta"]  = round(after_b, 4)
        ws_tune.append([today_s,"D-7","d7_alpha",before_a,state["calibration"]["d7_alpha"],round(state["calibration"]["d7_alpha"]-before_a,4),"Brier 기반 미세 조정"])
        ws_tune.append([today_s,"D-7","d7_beta", before_b,state["calibration"]["d7_beta"], round(state["calibration"]["d7_beta"]-before_b,4),"Brier 기반 미세 조정"])

    # 3-2) D-1: precision 목표로 d1_high 조정 (○의 '정확도' 관리)
    d1_ws = wb["D-1"]
    tp=fp=0
    preds=[]
    for row in d1_ws.iter_rows(min_row=2, values_only=True):
        if row[1] != today_s:
            continue
        p = row[5]; y = row[8]
        if p is None or y is None: 
            continue
        p = float(p); y=int(y)
        preds.append((p,y))
    if preds:
        before = float(state["threshold"]["d1_high"])
        # precision 평가
        # 현재 컷오프로 precision 계산
        def precision_at(th):
            sel = [(p,y) for (p,y) in preds if p >= th]
            if not sel:
                return 1.0  # 선택 없음이면 precision 정의상 1로 간주(보수적)
            tp = sum(1 for p,y in sel if y==1)
            return tp/len(sel)
        cur_prec = precision_at(before)
        after = before
        if cur_prec < D1_PRECISION_TARGET:
            after = min(D1_HIGH_MAX, before + D1_HIGH_STEP)
            reason = f"정밀도 {D1_PRECISION_TARGET:.0%} 목표 상향"
        elif cur_prec > D1_PRECISION_TARGET + 0.05:
            after = max(D1_HIGH_MIN, before - D1_HIGH_STEP)
            reason = f"정밀도 과도 → 컷오프 완화"
        else:
            reason = "유지"
        if after != before:
            state["threshold"]["d1_high"] = round(after, 3)
            ws_tune.append([today_s,"D-1","d1_high",before,state["threshold"]["d1_high"],round(state["threshold"]["d1_high"]-before,3),reason])

    wb.save(REPORT_XLSX)
    save_state(state)

# -------------------- Excel 생성 --------------------
def autofit(ws):
    dims: Dict[int, int] = {}
    for row in ws.rows:
        for cell in row:
            if cell.value is None:
                continue
            l = len(str(cell.value).encode("utf-8"))
            idx = cell.col_idx
            dims[idx] = max(dims.get(idx, 10), l+2)
    for idx, width in dims.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(60, width)

def build_workbook(dates: List[dt.date], rooms: List[Room], events_map: Dict[RoomKey, List[Event]], state: dict) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    summary_rows: List[List] = []

    for d in dates:
        ws = wb.create_sheet(d.strftime("%Y%m%d"))
        ws.append(["sector","building","out","early","p50","low","hi","potential","total"])

        agg: Dict[Tuple[str,str], Dict[str, float]] = {}
        detail_rows: List[List] = []

        h = max(1, (d - TODAY).days)  # 안전장치: 과거 입력 시에도 최소 D+1 처리
        _, _, high_for_h = horizon_params(h, state)
        borderline = float(state["threshold"]["borderline"])

        # 방별 계산
        for r in rooms:
            evs = events_map.get(r.key, [])
            flags = occupancy_flags(evs, d)
            p = p_room_out(evs, d, h, state)

            # potential 기호
            pot_sym = ""
            if not flags["out"] and not flags["stay"]:
                if p >= high_for_h:
                    pot_sym = "○"
                elif p >= borderline:
                    pot_sym = "△"

            total_c = 1 if (flags["out"] or pot_sym == "○") else 0
            out_cell = r.out_time.strftime("%H:%M") if flags["out"] else ""

            detail_rows.append([
                r.key.sector, r.key.building, r.key.room,
                out_cell, f"{p:.2f}", pot_sym, total_c
            ])

            k = (r.key.sector, r.key.building)
            if k not in agg:
                agg[k] = {"out":0, "early":0, "mu":0.0, "var":0.0, "pot":0, "tot":0}
            if flags["out"]:
                agg[k]["out"] += 1
                if r.out_time < dt.time(12,0):
                    agg[k]["early"] += 1
            # μ, σ²(확정/체류 반영한 p)
            agg[k]["mu"]  += p
            agg[k]["var"] += p*(1-p)
            if pot_sym == "○":
                agg[k]["pot"] += 1
            if total_c:
                agg[k]["tot"] += 1

        # Header: 정렬 sector, out DESC
        for (sector, building), v in sorted(agg.items(), key=lambda x: (x[0][0], -x[1]["out"])):
            mu = v["mu"]; sigma = math.sqrt(max(0.0, v["var"]))
            p50 = int(round(mu))
            low = int(max(0, math.floor(mu - 1.28*sigma)))
            hi  = int(math.ceil(mu + 1.28*sigma))
            ws.append([sector, building, int(v["out"]), int(v["early"]), p50, low, hi, int(v["pot"]), int(v["tot"])])
            summary_rows.append([d.strftime("%Y%m%d"), sector, building, int(v["out"]), int(v["early"]), p50, low, hi, int(v["pot"]), int(v["tot"])])

        # Detail: 정렬 sector, out유무, building, room
        ws.append([])
        ws.append(["sector","building","room","out","p_out","potential","total"])
        for row in sorted(detail_rows, key=lambda r: (r[0], -int(bool(r[3])), r[1], r[2])):
            ws.append(row)

        autofit(ws)

    # Summary 생성
    from itertools import groupby
    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(["date","sector","building","out","early","p50","low","hi","potential","total"])
    for date_str, grp in groupby(sorted(summary_rows, key=lambda x: x[0]), key=lambda x: x[0]):
        grp_list = list(grp)
        tot_out = sum(r[3] for r in grp_list)
        tot_early = sum(r[4] for r in grp_list)
        tot_p50 = sum(r[5] for r in grp_list)
        tot_low = sum(r[6] for r in grp_list)
        tot_hi = sum(r[7] for r in grp_list)
        tot_pot = sum(r[8] for r in grp_list)
        tot_tot = sum(r[9] for r in grp_list)
        ws_sum.append([date_str,"TOTAL","-", tot_out, tot_early, tot_p50, tot_low, tot_hi, tot_pot, tot_tot])
        r_idx = ws_sum.max_row
        for c in range(1, 11):
            cell = ws_sum.cell(r_idx, c)
            cell.fill = PatternFill("solid", fgColor="000000")
            cell.font = Font(color="FFFFFF", bold=True)
        for r in grp_list:
            ws_sum.append(r)
    autofit(ws_sum)

    return wb

# -------------------- 메인 --------------------
def main():
    parser = argparse.ArgumentParser(description="Forecasting v1.3 (fresh build)")
    parser.add_argument("--auto", action="store_true", help="D+1..D+8 자동 예측 후 종료")
    args = parser.parse_args()

    # data.toml 로드
    try:
        rooms = load_data_toml("data.toml")
    except Exception as e:
        print(f"[에러] data.toml 로드 실패: {e}")
        if not args.auto:
            # 수동 모드: 10분 대기(입력시 즉시 종료)
            import time, select
            print("\n[종료 대기] 아무 키나 누르면 종료. 입력이 없으면 10분 후 자동 종료.")
            start = time.time()
            while time.time() - start < 600:
                try:
                    r,_,_ = select.select([sys.stdin], [], [], 0.2)
                    if r:
                        _ = sys.stdin.readline()
                        break
                except Exception:
                    time.sleep(0.2)
            print("종료합니다.")
        sys.exit(2)

    # 날짜 결정
    if args.auto:
        dates = [TODAY + dt.timedelta(days=i) for i in range(1, 9)]  # D+1..D+8
    else:
        while True:
            raw = input("날짜 입력 (YYYYMMDD, 콤마/틸다 혼용, 예: 20251108,20251110~20251112): ").strip()
            try:
                dates = parse_dates(raw)
                break
            except Exception as e:
                print(f"[에러] 날짜 형식 오류: {e}")

    first_date = min(dates)
    out_xlsx = unique_filename(f"forecasting({first_date.strftime('%Y%m%d')})", ".xlsx")

    # ICS 다운로드/파싱
    ts = dt.datetime.now(SEOUL).strftime("%Y%m%d%H%M%S")
    ics_dir = ensure_ics_dir(ts)
    ics_map = download_ics(rooms, ts, ics_dir)
    events_map: Dict[RoomKey, List[Event]] = {rk: parse_ics(paths) for rk, paths in ics_map.items()}

    # report.xlsx 준비 + 오늘(D0) 리포트 업데이트 (actual/accuracy/tuning)
    state = load_state()
    report = open_or_create_report()
    update_actuals_and_accuracy_and_tune(report, TODAY, rooms, events_map, state)

    # 예측 + forecasting.xlsx 생성
    wb = build_workbook(dates, rooms, events_map, state)
    wb.save(out_xlsx)
    print(f"[완료] 엑셀 저장: {out_xlsx}")

    # report.xlsx에는 D-1/D-7 예측만 저장 (D-2~D-6 저장 안 함)
    rows_d1: List[List] = []
    rows_d7: List[List] = []
    run_date_s = TODAY.strftime("%Y%m%d")
    for d in dates:
        h = (d - TODAY).days
        if h not in (1,7):
            continue
        _, _, high = horizon_params(h, state)
        borderline = float(state["threshold"]["borderline"])
        target_s = d.strftime("%Y%m%d")
        for r in rooms:
            evs = events_map.get(r.key, [])
            p = p_room_out(evs, d, h, state)
            pred_label = 1 if p >= high else 0
            pot_sym = "○" if p >= high else ("△" if p >= borderline else "")
            row = [run_date_s, target_s, r.key.sector, r.key.building, r.key.room, f"{p:.3f}", pred_label, pot_sym, None, None]
            if h == 1:
                rows_d1.append(row)
            else:
                rows_d7.append(row)

    if rows_d1:
        append_prediction(report, "D-1", rows_d1)
    if rows_d7:
        append_prediction(report, "D-7", rows_d7)

    # --auto: CMD까지 종료
    if args.auto:
        print("[자동 모드 완료 — CMD 창을 닫습니다]")
        os._exit(0)

if __name__ == "__main__":
    main()
